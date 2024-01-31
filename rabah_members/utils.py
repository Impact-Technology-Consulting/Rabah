import csv
import json
import operator
import re
from datetime import datetime
from functools import reduce
from io import TextIOWrapper

from django.db.models import Q
from openpyxl import load_workbook

from .models import Member


def get_member(user, organisation_id):
    """"
    This is used to get a member or an organisation
    """
    member = Member.objects.filter(user=user, organisation_id=organisation_id).first()
    if not member:
        raise
    return member


def query_members(query, item):
    """
    this query list is used to filter item more of like a custom query the return the query set
    :param query:
    :param item:
    :return:
    """
    query_list = []
    query_list += query.split()
    query_list = sorted(query_list, key=lambda x: x[-1])
    query = reduce(
        operator.or_,
        (Q(user__email__icontains=x) |
         Q(user__first_name__icontains=x) |
         Q(user__last_name__icontains=x) |
         Q(user__first_name=[x]) for x in query_list)
    )
    object_list = item.filter(query).distinct()
    return object_list


class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.isoformat()
        return super().default(obj)


def excel_to_dict_list(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active

    data = []
    #  create custom header for check
    headers = []
    header_dictionary = []
    for cell in ws[1]:
        # if there is a cell, then it appends it to the header and the header dictionary
        headers.append(convert_string(cell.value))
        header_dictionary.append({convert_string(cell.value): cell.value})

    for row in ws.iter_rows(min_row=2):  # Assuming the data starts from the third row
        row_data = {}
        all_none = True
        for header, cell in zip(headers, row):
            # loop through the headers and the row accordingly
            row_data[header] = cell.value
            if cell.value is not None:
                all_none = False
        if all_none:
            break
        data.append(row_data)

    return data, header_dictionary


def csv_to_dict_list(csv_file):
    data = []

    reader = csv.reader(TextIOWrapper(csv_file, encoding='utf-8'))
    raw_headers = next(reader)  # Read the first row as headers
    headers = [convert_string(header) for header in raw_headers]  # Clean the headers using convert_string function
    #  create custom header for check
    header_dictionary = []
    for item in raw_headers:
        if item:
            header_dictionary.append({convert_string(item): item})
    for row in reader:
        row_data = {}
        for header, value in zip(headers, row):
            row_data[header] = value
        data.append(row_data)

    return data, header_dictionary


def convert_file_to_dictionary(file):
    """
    this select the files to be used
    :param file:
    :return:
    """
    if str(file).endswith(".csv"):
        data, header_dictionary = csv_to_dict_list(file)
    elif str(file).endswith(".xlsx") or str(file).endswith(".xls"):
        data, header_dictionary = excel_to_dict_list(file)
    else:
        raise ValueError("Unsupported file format")

    return data, header_dictionary


def convert_string(input_string):
    """
    this is used to convert  the header with unknwon letters
    :param input_string:
    :return:
    """
    if not input_string or input_string == "":
        return None
    if type(input_string) != str:
        return input_string

    lowercase_string = input_string.lower()
    cleaned_string = re.sub('[^a-z0-9]', '', lowercase_string)
    return cleaned_string
